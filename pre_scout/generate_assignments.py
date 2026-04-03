"""
generate_assignments.py
Pulls the match schedule from TBA and assigns scouters to teams across all matches.
Works with any number of scouters. Each scouter is assigned one robot per match,
balanced so total assignments are as equal as possible across the event.

Usage:
    python pre_scout/generate_assignments.py

Requirements:
    pip install requests openpyxl
"""

import os
import sys
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────
TBA_KEY = os.getenv("TBA_KEY2")

EVENT  = "2026mifli"  
OUTPUT = "scouter_assignments.xlsx"

# Edit these names before each event — add or remove as needed
SCOUTERS = [
    "Collin",
    "Joe",
    "Liz",
    "Peyton",
    "Ruthie"
]
# ─────────────────────────────────────────────────────────────────────────────

TBA_BASE = "https://www.thebluealliance.com/api/v3"
TBA_HDR  = {"X-TBA-Auth-Key": TBA_KEY}

ROBOTS_PER_MATCH = 6  # always 6 in FRC (3 red + 3 blue)


def tba_get(path):
    r = requests.get(f"{TBA_BASE}{path}", headers=TBA_HDR, timeout=15)
    r.raise_for_status()
    return r.json()


def fetch_schedule():
    print("Fetching match schedule from TBA...")
    try:
        matches = tba_get(f"/event/{EVENT}/matches/simple")
    except Exception as e:
        print(f"Fatal: Could not fetch schedule — {e}")
        sys.exit(1)

    qual_matches = [m for m in matches if m["comp_level"] == "qm"]
    qual_matches.sort(key=lambda m: m["match_number"])

    if not qual_matches:
        print("No qualification matches found. Schedule may not be released yet.")
        sys.exit(1)

    print(f"  Found {len(qual_matches)} qualification matches.")
    return qual_matches


def build_assignments(matches):
    """
    Assigns scouters to robots across all matches.
    Each match has 6 robots. With N scouters, each match is divided into
    6 slots assigned round-robin to scouters, always giving the next slot
    to whichever scouter currently has the lowest total load.
    """
    n = len(SCOUTERS)
    load = {s: 0 for s in SCOUTERS}

    # Each match produces 6 robot slots
    slot_labels = ["Red 1", "Red 2", "Red 3", "Blue 1", "Blue 2", "Blue 3"]

    # Per-scouter rows: list of dicts
    scouter_rows = {s: [] for s in SCOUTERS}
    # Master schedule rows
    schedule_rows = []

    for match in matches:
        match_num  = match["match_number"]
        red_teams  = [t.replace("frc", "") for t in match["alliances"]["red"]["team_keys"]]
        blue_teams = [t.replace("frc", "") for t in match["alliances"]["blue"]["team_keys"]]
        all_teams  = red_teams + blue_teams  # indices 0-5 map to slot_labels

        match_assignments = {}  # slot_label -> scouter

        for slot_idx, (label, team) in enumerate(zip(slot_labels, all_teams)):
            # Assign to the scouter with the lowest current load
            # Tie-break by scouter list order for consistency
            assigned = min(SCOUTERS, key=lambda s: (load[s], SCOUTERS.index(s)))
            load[assigned] += 1
            match_assignments[label] = (team, assigned)

            scouter_rows[assigned].append({
                "Match":    match_num,
                "Slot":     label,
                "Team":     team,
            })

        schedule_rows.append({
            "match_num":        match_num,
            "assignments":      match_assignments,
            "all_teams":        all_teams,
        })

    return schedule_rows, scouter_rows, load


def style_header(cell, bg="1565C0"):
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def data_cell(ws, row, col, value, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = thin_border()
    if fill:
        cell.fill = fill
    return cell


def build_xlsx(schedule_rows, scouter_rows, load):
    wb = Workbook()
    wb.remove(wb.active)

    alt       = PatternFill("solid", start_color="EEF2FF")
    red_fill  = PatternFill("solid", start_color="FFEBEE")
    blue_fill = PatternFill("solid", start_color="E3F2FD")

    slot_labels = ["Red 1", "Red 2", "Red 3", "Blue 1", "Blue 2", "Blue 3"]

    # ── Sheet 1: Full Schedule ────────────────────────────────────────────
    ws = wb.create_sheet("Full Schedule")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    # Headers: Match | Red1 Team | Red1 Scouter | Red2 ... | Blue3 Scouter
    headers = ["Match"]
    for label in slot_labels:
        headers.append(f"{label} Team")
        headers.append(f"{label} Scouter")

    col_widths = [8] + [10, 14] * 6

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        # Color red/blue headers
        if "Red" in h:
            style_header(cell, "C62828")
        elif "Blue" in h:
            style_header(cell, "1565C0")
        else:
            style_header(cell, "37474F")

    for row_idx, row_data in enumerate(schedule_rows, 2):
        fill = alt if row_idx % 2 == 0 else None
        data_cell(ws, row_idx, 1, row_data["match_num"], fill)
        col = 2
        for label in slot_labels:
            team, scouter = row_data["assignments"][label]
            slot_fill = red_fill if "Red" in label else blue_fill
            data_cell(ws, row_idx, col,     team,    slot_fill)
            data_cell(ws, row_idx, col + 1, scouter, slot_fill)
            col += 2

    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Sheet 2+: Per-Scouter Sheets ─────────────────────────────────────
    scouter_colors = [
        "1B5E20", "4A148C", "E65100", "006064", "BF360C", "1A237E",
        "880E4F", "0D47A1", "33691E", "4E342E",
    ]

    for i, scouter in enumerate(SCOUTERS):
        color = scouter_colors[i % len(scouter_colors)]
        ws_s  = wb.create_sheet(scouter)
        ws_s.freeze_panes = "A2"
        ws_s.row_dimensions[1].height = 28

        # Header
        s_headers   = ["Match", "Alliance Slot", "Team to Scout"]
        s_widths    = [8, 14, 16]

        for col, h in enumerate(s_headers, 1):
            cell = ws_s.cell(row=1, column=col, value=h)
            style_header(cell, color)

        rows = scouter_rows[scouter]
        for row_idx, r in enumerate(rows, 2):
            f = alt if row_idx % 2 == 0 else None
            slot_fill = red_fill if "Red" in r["Slot"] else blue_fill
            data_cell(ws_s, row_idx, 1, r["Match"], f)
            data_cell(ws_s, row_idx, 2, r["Slot"],  slot_fill)
            data_cell(ws_s, row_idx, 3, r["Team"],  slot_fill)

        for col, width in enumerate(s_widths, 1):
            ws_s.column_dimensions[get_column_letter(col)].width = width

        # Summary at the bottom
        total_row = len(rows) + 3
        ws_s.cell(row=total_row, column=1, value="Total Assignments:").font = Font(bold=True, name="Arial", size=10)
        ws_s.cell(row=total_row, column=2, value=load[scouter]).font = Font(bold=True, name="Arial", size=10)

    # ── Sheet: Load Summary ───────────────────────────────────────────────
    ws_l = wb.create_sheet("Load Summary")
    ws_l.row_dimensions[1].height = 28

    for col, h in enumerate(["Scouter", "Total Assignments"], 1):
        cell = ws_l.cell(row=1, column=col, value=h)
        style_header(cell, "37474F")

    for row_idx, scouter in enumerate(SCOUTERS, 2):
        f = alt if row_idx % 2 == 0 else None
        data_cell(ws_l, row_idx, 1, scouter,       f)
        data_cell(ws_l, row_idx, 2, load[scouter], f)

    ws_l.column_dimensions["A"].width = 20
    ws_l.column_dimensions["B"].width = 20

    wb.save(OUTPUT)


def main():
    if len(SCOUTERS) == 0:
        print("Error: SCOUTERS list is empty. Add scouter names to the config.")
        sys.exit(1)

    print(f"Scouters ({len(SCOUTERS)}): {', '.join(SCOUTERS)}")
    print(f"Event: {EVENT}")
    print()

    matches          = fetch_schedule()
    schedule_rows, scouter_rows, load = build_assignments(matches)

    print("\nAssignment load:")
    for s, count in load.items():
        print(f"  {s}: {count} robots")

    build_xlsx(schedule_rows, scouter_rows, load)

    print(f"\n✅ Saved: {OUTPUT}")
    print(f"   Sheets: Full Schedule | {' | '.join(SCOUTERS)} | Load Summary")


if __name__ == "__main__":
    main()