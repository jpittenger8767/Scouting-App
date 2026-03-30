"""
FRC Pre-Event Scouting Spreadsheet Generator
Pulls data from The Blue Alliance and Statbotics, outputs a formatted .xlsx file.

Requirements:
    pip install requests openpyxl statbotics

Usage:
    python generate_scouting_sheet.py
"""

import requests
import os
import statbotics
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────
TBA_KEY = os.getenv("TBA_KEY2")
EVENT    = "2026miken" 
OUTPUT   = f"{EVENT}_scouting.xlsx"
# ─────────────────────────────────────────────────────────────────────────────

TBA_BASE = "https://www.thebluealliance.com/api/v3"
TBA_HDR  = {"X-TBA-Auth-Key": TBA_KEY}


def tba_get(path):
    r = requests.get(f"{TBA_BASE}{path}", headers=TBA_HDR, timeout=15)
    r.raise_for_status()
    return r.json()


def fetch_data():
    print("Fetching team list from TBA...")
    teams = tba_get(f"/event/{EVENT}/teams")
    team_numbers = sorted(int(t["team_number"]) for t in teams)
    team_map = {t["team_number"]: t for t in teams}

    print("Fetching OPRs from TBA...")
    try:
        oprs_raw = tba_get(f"/event/{EVENT}/oprs")
        oprs  = oprs_raw.get("oprs",  {})
        dprs  = oprs_raw.get("dprs",  {})
        ccwms = oprs_raw.get("ccwms", {})
    except Exception:
        print("  OPRs not available yet (event hasn't started).")
        oprs = dprs = ccwms = {}

    print("Fetching rankings from TBA...")
    try:
        rankings_raw = tba_get(f"/event/{EVENT}/rankings")
        rank_list = rankings_raw.get("rankings", [])
        rankings = {}
        for r in rank_list:
            tn = int(r["team_key"].replace("frc", ""))
            rankings[tn] = {
                "rank":   r.get("rank"),
                "wins":   r.get("record", {}).get("wins"),
                "losses": r.get("record", {}).get("losses"),
                "ties":   r.get("record", {}).get("ties"),
                "rp":     r.get("extra_stats", [None])[0] if r.get("extra_stats") else None,
            }
    except Exception:
        print("  Rankings not available yet (event hasn't started).")
        rankings = {}

    print("Fetching 2026 EPA data from Statbotics...")
    stat_map = {}
    try:
        sb = statbotics.Statbotics()
        year = int("".join(filter(str.isdigit, EVENT)))

        for tn in team_numbers:
            try:
                s = sb.get_team_year(tn, year)
                if s and "epa" in s:
                    e = s["epa"]
                    
                    # Based on the exact 2026 JSON structure
                    breakdown = e.get("breakdown", {})
                    stats = e.get("stats", {})
                    
                    stat_map[tn] = {
                        "epa_start":        stats.get("start"),
                        "epa_pre_playoffs": stats.get("pre_champs"),       # Mapped from 'pre_champs'
                        "epa_end":          breakdown.get("total_points"), # The current total EPA
                        "auto_epa":         breakdown.get("auto_points"),  # Found the exact key!
                        "teleop_epa":       breakdown.get("teleop_points"),# Found the exact key!
                        "endgame_epa":      breakdown.get("endgame_points"),# Found the exact key!
                        "norm_epa":         e.get("norm"),                 # We already knew this worked
                    }
            except Exception:
                continue
    except Exception as e:
        print(f"  Statbotics Connection Error: {e}")

    return team_numbers, team_map, oprs, dprs, ccwms, rankings, stat_map


def style_header_cell(cell, bg_hex="1565C0"):
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg_hex)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def build_sheet(wb, sheet_name, headers, rows, col_widths, header_color="1565C0"):
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        style_header_cell(cell, header_color)

    alt_fill = PatternFill("solid", start_color="EEF2FF")

    for row_idx, row_data in enumerate(rows, 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border()
            if fill:
                cell.fill = fill

    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    return ws


def fmt(val, decimals=2):
    if val is None:
        return "N/A"
    try:
        return round(float(val), decimals)
    except (TypeError, ValueError):
        return "N/A"


def main():
    team_numbers, team_map, oprs, dprs, ccwms, rankings, stat_map = fetch_data()

    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Overview ─────────────────────────────────────────────────────
    headers_main = [
        "Team #", "Nickname", "City / State",
        "Rank", "W", "L", "T", "RP",
        "OPR", "DPR", "CCWM",
        "EPA Start", "EPA Pre-Elim", "EPA End", "Norm EPA",
        "Auto EPA", "Teleop EPA", "Endgame EPA",
    ]
    col_widths_main = [9, 22, 20, 7, 6, 6, 6, 8, 9, 9, 9, 11, 13, 10, 11, 11, 13, 14]

    rows_main = []
    for tn in team_numbers:
        key  = f"frc{tn}"
        info = team_map.get(tn, {})
        rank = rankings.get(tn, {})
        stat = stat_map.get(tn, {})
        city_state = ", ".join(filter(None, [info.get("city"), info.get("state_prov")]))

        rows_main.append([
            tn,
            info.get("nickname", "N/A"),
            city_state or "N/A",
            rank.get("rank",   "N/A"),
            rank.get("wins",   "N/A"),
            rank.get("losses", "N/A"),
            rank.get("ties",   "N/A"),
            fmt(rank.get("rp"), 1),
            fmt(oprs.get(key)),
            fmt(dprs.get(key)),
            fmt(ccwms.get(key)),
            fmt(stat.get("epa_start")),
            fmt(stat.get("epa_pre_playoffs")),
            fmt(stat.get("epa_end")),
            fmt(stat.get("norm_epa")),
            fmt(stat.get("auto_epa")),
            fmt(stat.get("teleop_epa")),
            fmt(stat.get("endgame_epa")),
        ])

    rows_main.sort(key=lambda r: (r[3] == "N/A", r[3] if r[3] != "N/A" else 9999))
    build_sheet(wb, "Overview", headers_main, rows_main, col_widths_main, "1565C0")

    # ── Sheet 2: EPA (Statbotics) ─────────────────────────────────────────────
    headers_sb = ["Team #", "Nickname", "EPA Start", "EPA Prelim", "EPA End", "Auto EPA", "Teleop EPA", "Endgame EPA", "Norm EPA"]
    col_widths_sb = [9, 22, 10, 10, 10, 10, 10, 10, 10]

    rows_sb = []
    for tn in team_numbers:
        info  = team_map.get(tn, {})
        stats = stat_map.get(tn, {})
        rows_sb.append([
            tn,
            info.get("nickname", "N/A"),
            fmt(stats.get("epa_start")),
            fmt(stats.get("epa_pre_playoffs")),
            fmt(stats.get("epa_end")),
            fmt(stats.get("auto_epa")),
            fmt(stats.get("teleop_epa")),
            fmt(stats.get("endgame_epa")),
            fmt(stats.get("norm_epa")),
        ])
    
    # Sort by EPA End (Column index 4) descending
    rows_sb.sort(key=lambda r: (r[4] == "N/A", -r[4] if r[4] != "N/A" else 0))
    build_sheet(wb, "EPA (Statbotics)", headers_sb, rows_sb, col_widths_sb, "1F8E3C")

    # ── Sheet 3: OPR/DPR/CCWM ────────────────────────────────────────────────
    headers_opr = ["Team #", "Nickname", "OPR", "DPR", "CCWM"]
    col_widths_opr = [9, 22, 10, 10, 10]

    rows_opr = []
    for tn in team_numbers:
        key  = f"frc{tn}"
        info = team_map.get(tn, {})
        rows_opr.append([
            tn,
            info.get("nickname", "N/A"),
            fmt(oprs.get(key)),
            fmt(dprs.get(key)),
            fmt(ccwms.get(key)),
        ])
    rows_opr.sort(key=lambda r: (r[2] == "N/A", -r[2] if r[2] != "N/A" else 0))
    build_sheet(wb, "OPR (TBA)", headers_opr, rows_opr, col_widths_opr, "4A148C")

    # ── Sheet 4: Rankings ─────────────────────────────────────────────────────
    headers_rank = ["Rank", "Team #", "Nickname", "W", "L", "T", "RP"]
    col_widths_rank = [7, 9, 22, 6, 6, 6, 8]

    rows_rank = []
    for tn in team_numbers:
        info = team_map.get(tn, {})
        rank = rankings.get(tn, {})
        rows_rank.append([
            rank.get("rank",   "N/A"),
            tn,
            info.get("nickname", "N/A"),
            rank.get("wins",   "N/A"),
            rank.get("losses", "N/A"),
            rank.get("ties",   "N/A"),
            fmt(rank.get("rp"), 1),
        ])
    rows_rank.sort(key=lambda r: (r[0] == "N/A", r[0] if r[0] != "N/A" else 9999))
    build_sheet(wb, "Rankings (TBA)", headers_rank, rows_rank, col_widths_rank, "B71C1C")

    # ── Sheet 5: Pick List ────────────────────────────────────────────────────────
headers_pick = ["Pick Order", "Team #", "Nickname", "OPR", "DPR", "CCWM", "DNP (x = exclude)"]
col_widths_pick = [10, 9, 22, 10, 10, 10, 18]

rows_pick = []
for tn in team_numbers:
    key  = f"frc{tn}"
    info = team_map.get(tn, {})
    opr_val = oprs.get(key)
    rows_pick.append([
        None,  # Pick order filled in after sorting
        tn,
        info.get("nickname", "N/A"),
        fmt(opr_val),
        fmt(dprs.get(key)),
        fmt(ccwms.get(key)),
        "",  # DNP column, manually filled
    ])

# Sort by OPR descending
rows_pick.sort(key=lambda r: (r[3] == "N/A", -r[3] if r[3] != "N/A" else 0))

# Fill in pick order after sorting
for i, row in enumerate(rows_pick, 1):
    row[0] = i

build_sheet(wb, "2nd Pick List", headers_pick, rows_pick, col_widths_pick, "E65100")

    wb.save(OUTPUT)
    print(f"\n✅ Saved: {OUTPUT}")
    print(f"   Sheets: Overview | EPA (Statbotics) | OPR (TBA) | Rankings (TBA)")
    print(f"   Teams:  {len(team_numbers)}")


if __name__ == "__main__":
    main()
