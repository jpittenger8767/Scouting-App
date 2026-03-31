"""
merge_scouting.py
Merges all scouting CSVs in a folder into a single formatted .xlsx file.

Usage:
    python merge_scouting.py                  # merges CSVs in current directory
    python merge_scouting.py path/to/csvs     # merges CSVs in specified folder
"""

import sys
import glob
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "scouting_merged.xlsx"

EXPECTED_COLS = [
    "Match", "Team", "Alliance", "Scouter",
    "Auto Fuel", "Auto Mobility", "Auto Climb",
    "Teleop Fuel", "Field Zone", "Endgame Climb",
    "Defense Rating", "Driver Rating", "Played Defense",
    "Notes", "Timestamp",
]

def merge_csvs(folder="."):
    pattern = os.path.join(folder, "*.csv")
    files   = glob.glob(pattern)
    if not files:
        print(f"No CSV files found in '{folder}'.")
        sys.exit(1)

    dfs = []
    for f in files:
        try:
            df = pd.read_csv(f)
            df["_source"] = os.path.basename(f)
            dfs.append(df)
            print(f"  Loaded: {os.path.basename(f)} ({len(df)} rows)")
        except Exception as e:
            print(f"  Skipping {f}: {e}")

    merged = pd.concat(dfs, ignore_index=True)
    merged.sort_values(["Match", "Team"], inplace=True)
    merged.reset_index(drop=True, inplace=True)
    return merged


def style_header(cell, bg="1565C0"):
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def write_sheet(wb, name, df, col_order, col_widths, header_color="1565C0"):
    ws = wb.create_sheet(name)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    display_cols = [c for c in col_order if c in df.columns]
    for col_idx, col_name in enumerate(display_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        style_header(cell, header_color)

    alt = PatternFill("solid", start_color="EEF2FF")
    for row_idx, (_, row) in enumerate(df[display_cols].iterrows(), 2):
        fill = alt if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = thin_border()
            if fill:
                cell.fill = fill

    for col_idx, width in enumerate(col_widths[:len(display_cols)], 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def main():
    folder = sys.argv[1] if len(sys.argv) > 1 else "."
    print(f"Scanning for CSVs in: {os.path.abspath(folder)}")
    df = merge_csvs(folder)
    print(f"\nTotal entries merged: {len(df)}")

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: All Data ─────────────────────────────────────────────────
    all_cols   = EXPECTED_COLS + ["_source"]
    all_widths = [8,9,9,14,10,14,12,12,12,14,14,13,14,35,22,18]
    write_sheet(wb, "All Entries", df, all_cols, all_widths, "1565C0")

    # ── Sheet 2: By Team ──────────────────────────────────────────────────
    team_cols   = ["Team","Match","Alliance","Auto Fuel","Auto Mobility","Auto Climb",
                   "Teleop Fuel","Field Zone","Endgame Climb",
                   "Defense Rating","Driver Rating","Played Defense","Notes"]
    team_widths = [9,8,9,10,14,12,12,12,14,14,13,14,35]
    df_team = df.sort_values(["Team","Match"])
    write_sheet(wb, "By Team", df_team, team_cols, team_widths, "1F8E3C")

    # ── Sheet 3: Summary ──────────────────────────────────────────────────
    numeric_df = df.copy()
    for col in ["Auto Fuel","Teleop Fuel","Defense Rating","Driver Rating"]:
        numeric_df[col] = pd.to_numeric(numeric_df[col], errors="coerce")

    climb_pts = {"None":0,"L1":10,"L2":20,"L3":30}
    numeric_df["Endgame Pts"] = numeric_df["Endgame Climb"].map(climb_pts).fillna(0)

    summary = numeric_df.groupby("Team").agg(
        Matches       = ("Match",         "count"),
        Avg_Auto_Fuel = ("Auto Fuel",      "mean"),
        Avg_Tel_Fuel  = ("Teleop Fuel",    "mean"),
        Avg_Defense   = ("Defense Rating", "mean"),
        Avg_Driver    = ("Driver Rating",  "mean"),
        Avg_Endgame   = ("Endgame Pts",    "mean"),
    ).reset_index()

    summary["Avg_Auto_Fuel"] = summary["Avg_Auto_Fuel"].round(1)
    summary["Avg_Tel_Fuel"]  = summary["Avg_Tel_Fuel"].round(1)
    summary["Avg_Defense"]   = summary["Avg_Defense"].round(2)
    summary["Avg_Driver"]    = summary["Avg_Driver"].round(2)
    summary["Avg_Endgame"]   = summary["Avg_Endgame"].round(1)
    summary["Avg_Total_Fuel"]= (summary["Avg_Auto_Fuel"] + summary["Avg_Tel_Fuel"]).round(1)
    summary.sort_values("Avg_Total_Fuel", ascending=False, inplace=True)

    summary.rename(columns={
        "Avg_Auto_Fuel":  "Avg Auto Fuel",
        "Avg_Tel_Fuel":   "Avg Teleop Fuel",
        "Avg_Total_Fuel": "Avg Total Fuel",
        "Avg_Defense":    "Avg Defense",
        "Avg_Driver":     "Avg Driver Skill",
        "Avg_Endgame":    "Avg Endgame Pts",
    }, inplace=True)

    sum_cols   = ["Team","Matches","Avg Auto Fuel","Avg Teleop Fuel","Avg Total Fuel",
                  "Avg Endgame Pts","Avg Defense","Avg Driver Skill"]
    sum_widths = [9,9,14,16,15,15,13,16]
    write_sheet(wb, "Summary", summary, sum_cols, sum_widths, "4A148C")

    wb.save(OUTPUT)
    print(f"\n✅ Saved: {OUTPUT}")
    print(f"   Sheets: All Entries | By Team | Summary")


if __name__ == "__main__":
    main()